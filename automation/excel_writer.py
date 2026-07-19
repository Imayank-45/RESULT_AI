"""
excel_writer.py
---------------
Takes the list of parsed result dicts and writes a formatted .xlsx file:

    S.NO | ENROLLMENT | <subject columns, first-seen order> | RESULT | SGPA | CGPA

Formatting:
    - Bold header row with borders
    - Auto-fit column widths
    - RESULT text colored green for PASS, red for FAIL
    - Header row frozen
"""

from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PASS_FONT = Font(color="008000", bold=True)
FAIL_FONT = Font(color="FF0000", bold=True)


def _collect_subject_columns(records: List[dict]) -> List[str]:
    """First-seen order of subject codes across all records."""
    seen = []
    for rec in records:
        for code in rec.get("subjects", {}):
            if code not in seen:
                seen.append(code)
    return seen


def write_excel(records: List[dict], out_path: str) -> None:
    subject_columns = _collect_subject_columns(records)
    headers = ["S.NO", "ENROLLMENT", "NAME"] + subject_columns + ["RESULT", "SGPA", "CGPA"]

    wb = Workbook()
    ws = wb.active
    ws.title = "RGPV Results"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, rec in enumerate(records, start=1):
        if rec.get("not_found"):
            row = [i, rec.get("enrollment", ""), rec.get("name", "")] + [""] * (len(subject_columns) + 3)
            ws.append(row)
            continue

        result_status = rec.get("result", "UNKNOWN")
        result_display = result_status
        if result_status == "FAIL":
            fail_subjects = rec.get("fail_subjects") or []
            if fail_subjects:
                cleaned = []
                for s in fail_subjects:
                    clean_code = s.split("-")[0].strip() if "-" in s else s.strip()
                    if clean_code not in cleaned:
                        cleaned.append(clean_code)
                result_display = f"Fail in {', '.join(cleaned)}"

        row = [i, rec.get("enrollment", ""), rec.get("name", "")]
        for code in subject_columns:
            row.append(rec.get("subjects", {}).get(code, ""))
        row += [result_display, rec.get("sgpa", ""), rec.get("cgpa", "")]
        ws.append(row)

        result_col_idx = 3 + len(subject_columns) + 1
        result_cell = ws.cell(row=ws.max_row, column=result_col_idx)
        if result_status == "PASS":
            result_cell.font = PASS_FONT
        elif result_status == "FAIL":
            result_cell.font = FAIL_FONT

    # borders on all data cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = THIN_BORDER

    # auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"
    wb.save(out_path)
